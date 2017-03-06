import openpyxl
wb = openpyxl.Workbook()
print(wb.get_sheet_names())

wb.create_sheet()
#The create_sheet() method returns a new Worksheet object.
#Optionally, the index and name of the new sheet can be 
#specified with the index and title keyword argument. 
wb.create_sheet(index=0, title='First Sheet')

print(wb.get_sheet_names())

wb.create_sheet(index=2, title='Middle Sheet')
