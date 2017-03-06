import openpyxl

wb=openpyxl.load_workbook('example.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')
#The book uses sheet.get_highest_row(), but it has
#been deprecated as of version 2.3.someting. 
#The modern equivalent is sheet.max_row and it
#returns an int value. It's not a function. 
#But anyway, what this method does is tell us
#what is the number of rows and columns of the 
#ws object. 
print(str(sheet.max_row))

print(str(sheet.max_column))

