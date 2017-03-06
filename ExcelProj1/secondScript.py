import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')
#Cells can be referred by coordinates
print(sheet['A1'])
#The cell object calls a method to get its value
print(sheet['A1'].value)
#Here, we store the cell in a variable
c = sheet['B1']
#and so we can call on its value
print(c.value)
#its row and colum
print('Row ' + str(c.row) + ', Column '+c.column + ' is '+c.value)
#or its coordinate 
print('Cell ' + c.coordinate + ' is ' + c.value)

print(sheet['C1'].value)

#We can also call a specific cell by passing its 
#row and column in number form.
print("row 1 column 1:")
print(sheet.cell(row=1, column=2))

print(sheet.cell(row=1, column=2).value)

print("for statement:")

for i in range(1, 8, 2):
	print(i, sheet.cell(row=i, column=2).value)