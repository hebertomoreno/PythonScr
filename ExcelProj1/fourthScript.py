import openpyxl
#The book says these methods should be called from
#openpyxl.cell, but digging around a bit I found
#that they should be imported from openpyxl.utils
from openpyxl.utils import get_column_letter, column_index_from_string
#This function returns the letter for the column number
#it's given. 
colle = get_column_letter(1)

print(colle)

colle = get_column_letter(27)

print(colle)

colle = get_column_letter(900)

print(colle)

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

print(get_column_letter(sheet.max_column))
#This other function does the opposite. Given a letter,
#it returns the column number. 
colin = column_index_from_string('A')

print(colin)

colin = column_index_from_string('P')

print(colin)