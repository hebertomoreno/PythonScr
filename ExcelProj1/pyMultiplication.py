#! Python 3
#! pyMultiplication.py - Receives and arg from the 
#! command line and makes a table in excel with
#! its multiplication table. 

import openpyxl, sys
from openpyxl.styles import Font

if len(sys.argv) != 2:
	raise Exception('There must be a number after the execution statement.\n         Example: python pyMultiplication.py 10')
else:
	# Get the argument and convert it to int
	multNum = sys.argv[1]
	multNum = int(multNum)

	# Create a new workbook
	wb = openpyxl.Workbook()

	# Create the bold font
	boldFont = Font(bold=True)

	# Get the active sheet name
	sheet = wb.get_active_sheet()

	# Print column and row headers
	for headNum in range (1, multNum + 1):
		sheet.cell(row=1, column=headNum+1).font=boldFont
		sheet.cell(row=headNum+1, column = 1).font=boldFont
		sheet.cell(row=1, column=headNum+1).value = headNum;
		sheet.cell(row=headNum+1, column = 1).value = headNum;

	# Print the multiplied values
	for rowNum in range (2, sheet.max_row+1):
		for colNum in range(2, sheet.max_column+1):
			vertNum = sheet.cell(row=1, column=colNum).value
			horNum = sheet.cell(row=rowNum, column=1).value
			sheet.cell(row=rowNum, column=colNum).value = vertNum * horNum
	wb.save('multTable.xlsx')